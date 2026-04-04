#!/usr/bin/env python3
"""
Import Maharashtra Yojana (MJPJY / Ayushman Bharat) tariff from Excel to Supabase.
Usage: python3 scripts/import-maharashtra-yojana-tariff.py
"""

import openpyxl
import requests
import json
import sys
import os

SUPABASE_URL = "https://xvkxccqaopbnkvwgyfjv.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Inh2a3hjY3Fhb3Bibmt2d2d5Zmp2Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3NDc4MjMwMTIsImV4cCI6MjA2MzM5OTAxMn0.z9UkKHDm4RPMs_2IIzEPEYzd3-sbQSF6XpxaQg3vZhU"

EXCEL_PATH = os.path.expanduser("~/Downloads/NEW RET LIST MAHARASHTRA YOGNA.xlsx")

HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=minimal"
}

def safe_str(val):
    if val is None:
        return None
    return str(val).strip()

def safe_float(val):
    if val is None:
        return 0
    try:
        return float(str(val).replace(',', '').strip())
    except (ValueError, TypeError):
        return 0

def batch_insert(table, rows, batch_size=500):
    """Insert rows in batches."""
    total = len(rows)
    inserted = 0
    for i in range(0, total, batch_size):
        batch = rows[i:i + batch_size]
        resp = requests.post(
            f"{SUPABASE_URL}/rest/v1/{table}",
            headers=HEADERS,
            json=batch
        )
        if resp.status_code not in (200, 201):
            print(f"  ERROR inserting batch {i//batch_size + 1} into {table}: {resp.status_code} {resp.text[:200]}")
            return False
        inserted += len(batch)
        print(f"  {table}: {inserted}/{total} rows inserted")
    return True

def main():
    print(f"Loading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    # 1. Procedure sheet (main tariff)
    print("\n=== 1. Importing Procedures (main tariff) ===")
    ws = wb['Procedure sheet']
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not row[0]:  # skip empty rows
            continue
        rows.append({
            "procedure_code": safe_str(row[0]),
            "specialty": safe_str(row[1]),
            "specialty_code": safe_str(row[2]),
            "package_code": safe_str(row[3]),
            "package_name": safe_str(row[5]),
            "procedure_name": safe_str(row[6]),
            "tier3_rate": safe_float(row[7]),
            "implant_criteria": safe_str(row[8]) or 'N',
            "stratification_criteria": safe_str(row[9]) or 'N',
            "multiple_procedures": safe_str(row[10]) or 'No',
            "special_conditions": safe_str(row[11]) or 'N',
            "reservation_public": safe_str(row[12]) or 'N',
            "reservation_tertiary": safe_str(row[13]) or 'No',
            "level_of_care": safe_str(row[14]),
            "los": safe_str(row[15]),
            "auto_approved": safe_str(row[16]) or 'N',
            "mandatory_docs_preauth": safe_str(row[17]),
            "mandatory_docs_claim": safe_str(row[18]),
            "procedure_label": safe_str(row[19]),
            "special_condition_popup": safe_str(row[20]) or 'N',
            "special_conditions_rule": safe_str(row[21]) or 'N',
            "enhancement_applicable": safe_str(row[22]) or 'N',
            "medical_or_surgical": safe_str(row[23]),
            "day_care_procedure": safe_str(row[24]) or 'N',
            "reserved_procedure": safe_str(row[25]) if len(row) > 25 else None,
        })
    print(f"  Found {len(rows)} procedures")
    batch_insert("yojana_mh_procedures", rows)

    # 2. Implant Master
    print("\n=== 2. Importing Implant Master ===")
    ws = wb['Implant Master']
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        rows.append({
            "implant_code": safe_str(row[0]),
            "specialty": safe_str(row[1]),
            "implant_name": safe_str(row[2]),
            "procedure_code": safe_str(row[3]),
            "max_multiplier": safe_str(row[4]),
            "implant_price": safe_str(row[5]),
            "remarks": safe_str(row[6]) if len(row) > 6 else None,
        })
    print(f"  Found {len(rows)} implants")
    batch_insert("yojana_mh_implants", rows)

    # 3. Implant vs Procedure mapping
    print("\n=== 3. Importing Implant-Procedure mapping ===")
    ws = wb['Implant Vs Procedure']
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        rows.append({
            "implant_code": safe_str(row[0]),
            "procedure_code": safe_str(row[1]),
        })
    print(f"  Found {len(rows)} mappings")
    batch_insert("yojana_mh_implant_procedure_map", rows)

    # 4. Stratification Master
    print("\n=== 4. Importing Stratification Master ===")
    ws = wb['Stratification Master']
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        rows.append({
            "stratification_code": safe_str(row[0]),
            "stratification_options": safe_str(row[1]),
            "rule": safe_str(row[2]),
            "stratification_detail_code": safe_str(row[3]),
            "stratification_details": safe_str(row[4]),
            "stratification_detail_options": safe_str(row[5]),
            "override_procedure_price": safe_str(row[6]) if len(row) > 6 else None,
        })
    print(f"  Found {len(rows)} stratifications")
    batch_insert("yojana_mh_stratification", rows)

    # 5. Stratification vs Procedure mapping
    print("\n=== 5. Importing Stratification-Procedure mapping ===")
    ws = wb['Stratification Vs Procedure']
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        rows.append({
            "procedure_code": safe_str(row[0]),
            "stratification_code": safe_str(row[1]),
        })
    print(f"  Found {len(rows)} mappings")
    batch_insert("yojana_mh_stratification_procedure_map", rows)

    # 6. Special Condition Rules
    print("\n=== 6. Importing Special Condition Rules ===")
    ws = wb['Procedure Vs. SP rule mapping']
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        rows.append({
            "procedure_code": safe_str(row[0]),
            "rule_description": safe_str(row[1]),
        })
    print(f"  Found {len(rows)} rules")
    batch_insert("yojana_mh_special_conditions", rows)

    # 7. Add-On to Primary
    print("\n=== 7. Importing Add-On to Primary mappings ===")
    ws = wb['Add- On to Primary']
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        rows.append({
            "addon_procedure_code": safe_str(row[0]),
            "primary_procedure_code": safe_str(row[1]),
            "remarks": safe_str(row[2]) if len(row) > 2 else None,
        })
    print(f"  Found {len(rows)} add-on mappings")
    batch_insert("yojana_mh_addon_primary", rows)

    # 8. Add-On Specialty
    print("\n=== 8. Importing Add-On Specialty mappings ===")
    ws = wb['Add on Speciality']
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        rows.append({
            "procedure_code": safe_str(row[0]),
            "specialty_code": safe_str(row[1]),
        })
    print(f"  Found {len(rows)} specialty mappings")
    batch_insert("yojana_mh_addon_specialty", rows)

    # 9. Pop-Up Conditions
    print("\n=== 9. Importing Pop-Up Conditions ===")
    ws = wb['Procedure Vs. SP POP up mapping']
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        rows.append({
            "procedure_code": safe_str(row[0]),
            "popup_description": safe_str(row[1]),
            "stage": safe_str(row[2]) if len(row) > 2 else None,
            "step": safe_str(row[3]) if len(row) > 3 else None,
        })
    print(f"  Found {len(rows)} popup conditions")
    batch_insert("yojana_mh_popup_conditions", rows)

    # 10. Follow-Up to Procedure
    print("\n=== 10. Importing Follow-Up mappings ===")
    ws = wb['Follow Up to Procedure']
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        rows.append({
            "procedure_code": safe_str(row[0]),
            "followup_code": safe_str(row[1]) if len(row) > 1 else None,
            "remarks": safe_str(row[2]) if len(row) > 2 else None,
        })
    print(f"  Found {len(rows)} follow-up mappings")
    batch_insert("yojana_mh_followup_procedure", rows)

    # 11. Investigation Master
    print("\n=== 11. Importing Investigation Master ===")
    ws = wb['Investigation master']
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        rows.append({
            "investigation_code": safe_str(row[0]),
            "investigation_name": safe_str(row[1]),
        })
    print(f"  Found {len(rows)} investigations")
    batch_insert("yojana_mh_investigations", rows)

    # 12. Investigation vs Procedure mapping
    print("\n=== 12. Importing Investigation-Procedure mapping ===")
    ws = wb['Investigation Vs.procedure mapp']
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        rows.append({
            "investigation_code": safe_str(row[0]) if len(row) > 0 else None,
            "procedure_code": safe_str(row[1]) if len(row) > 1 else None,
        })
    print(f"  Found {len(rows)} investigation-procedure mappings")
    batch_insert("yojana_mh_investigation_procedure_map", rows)

    print("\n✅ Maharashtra Yojana tariff import complete!")
    wb.close()

if __name__ == "__main__":
    main()
